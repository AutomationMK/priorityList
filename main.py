import os
import shutil
import glob
import sys
import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Protection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.cell import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from downloadReports import download_reports


# check if the os is windows or linux and handle the filesystem based on that info
os_type = sys.platform
if os_type == "win32":
    priority_list_path = glob.glob(
        "L:\\Quotes\\Back Order Priority List\\priorityList*.xlsx"
    )
    quote_scan_folder = "L:\\Quotes\\Back Order Priority List\\"
elif os_type == "linux":
    priority_list_path = glob.glob(
        "/mnt/l/Quotes/Back Order Priority List/priorityList*.xlsx"
    )
    quote_scan_folder = "/mnt/l/Quotes/Back Order Priority List/"
else:
    print(f"Error: this program is not designed for os type {os_type}")
    exit(1)

# check if the correct amount of excel files in the back order priority list folder is 1
# if it is not 1 then it will report an error and end the program
if len(priority_list_path) > 1:
    print("Error: you cannot have more than one excel file in this folder")
    exit(1)
elif len(priority_list_path) == 0:
    print("Warning: No files found in the quote scans folder")
    print("The older priorityList file will be used instead if that is available")
    print("If this was an accident find a backup version of the file if it exists")
    print(
        "Then you can re-run the program to generate a correct file with all of the newer ship dates"
    )
else:
    # move the priorityList file to the local folder to be used by the program
    source = os.path.abspath(priority_list_path[0])
    dest = os.path.relpath("./priorityList.xlsx")
    try:
        shutil.move(source, dest)
    except FileNotFoundError:
        print(f"Error: Source file {source} not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# open E2 and download the needed reports
try:
    download_reports()
except Exception as e:
    print(e)
    exit()

data_folder = os.path.abspath(".") + "/csv_files"
# check if excel file location for excel file used is available
originalFile = "priorityList.xlsx"
# file location for Job information (assuming it is local)
scheduleLoc = f"{data_folder}/Job Schedule - Detail Report.csv"
# file location for PO information (assuming it is local)
poLoc = f"{data_folder}/Purchase Order Summary - Detail Report.csv"
# file location for order entry information (assuming it is local)
orderLoc = f"{data_folder}/Order Entry Summary - Detail Report.csv"
# list of Job Schedule header values
oldJobNum = "JobNumber"
oldPri = "Priority"
oldCust = "CustomerCode"
oldPartNum = "PartNumber"
oldRelType = "RelType"
oldCustQty = "QtyOrdered"
oldDueDate = "DueDate"
oldQtyShip = "QtyShipped"
oldShip = "ShipDate"
oldQtyReady = "QtyReady2Ship"
oldWorkCen = "WorkCenter"
oldJobNotes = "JobNotes"
oldRev = "Revision"
oldMasterJob = "MasterJobNo"
# list of PO header values
oldPoNum = "PONum"
oldPoVen = "Vendor"
oldPoDate = "Date"
oldPoPartNum = "PartNo"
oldPoPartDesc = "PartDesc"
oldPoGL = "GLAcct"
oldPoTotOrd = "TotalQtyOrdered"
oldPoPrice = "UnitPrice"
oldPoLine = "LineTotal"
oldPoDueDate = "DueDate"
oldPoJobNum = "JobNo"
oldPoQtyOrd = "QtyOrdered"
oldPoQtyRec = "QtyReceived"
oldPoQtyCanc = "QtyCanceled"
oldPoQtyRej = "QtyRejected"
oldPoJobs = "JobNumbers"
# list of order entry header values
oldOrdNum = "OrderNumber"
oldOrdDate = "OrderDate"
oldOrdPo = "PONumber"
oldOrdSales = "SalesID"
oldWorkCode = "WorkCode"
oldPartDesc = "PartDescription"
oldOrdQty = "Quantity"
oldOrdDisc = "Disc"
oldOrdTotal = "Total"
# list of new header values
jobNum = "Job Num"
ordDate = "Order Date"
cust = "Customer Code"
custPo = "Cust PO Num"
partNum = "Part Num"
relType = "Release Type"
custQty = "Qty Running"
dueDate = "Due Date"
workCode = "Work Code"
workCen = "Work Center"
estShipPrev2 = "Est Finish Prev 2"
estShipPrev = "Est Finish Prev"
estShip = "Est Finish"
notes = "Notes"
colePo = "Cole PO Number"
poVen = "PO Vendor"
poDate = "Cole PO Date"
poPartNum = "PO Part Num"
poPartDesc = "PO Part Desc"
totPoQty = "Total Qty Ordered"
poDueDate = "PO Due Date"
poQty = "Qty Ordered"
poQtyRec = "Qty Received"
poQtyRej = "Qty Rejected"
poLink = "PO Link"
# get the file timestamp for the Job Schedule Report CSV
c_time = os.path.getctime(scheduleLoc)
# convert the timestamp to a datetime object
version = datetime.datetime.fromtimestamp(c_time)
# add the timestamp to the file name for the finished excel file
newFile = f"priorityList-{version.strftime('%b-%d-%y %I-%M-%S %p')}.xlsx"

# list of previous excel file headers to remove
dropOldCol = [
    ordDate,
    cust,
    custPo,
    partNum,
    relType,
    custQty,
    dueDate,
    workCen,
    poLink,
]
# lists of Job Schedule headers to remove
dropSchedCol = [
    oldPri,
    oldQtyShip,
    oldShip,
    oldQtyReady,
    oldRev,
    oldMasterJob,
    oldJobNotes,
]
# list of PO headers to remove
dropPoCol = [
    oldPoGL,
    oldPoPrice,
    oldPoLine,
    oldPoQtyCanc,
    oldPoJobs,
]
# list of order entry headers to remove
dropOrdCol = [
    oldOrdNum,
    oldCust,
    oldOrdPo,
    oldOrdSales,
    oldPartDesc,
    oldOrdQty,
    oldPoPrice,
    oldOrdDisc,
    oldOrdTotal,
]
# read Job Schedule csv file to be place in a data frame
df = pd.read_csv(
    scheduleLoc,
    index_col=oldJobNum,
    dtype="string",
    parse_dates=[oldDueDate],
    date_format="%m/%d/%Y",
)
podf = pd.read_csv(
    poLoc,
    index_col=oldPoJobNum,
    dtype="string",
    parse_dates=[oldPoDate, oldPoDueDate],
    date_format="%m/%d/%Y",
)
orderdf = pd.read_csv(
    orderLoc,
    index_col=oldJobNum,
    dtype="string",
    parse_dates=[oldOrdDate],
    date_format="%m/%d/%Y",
)
# remove unused columns
df = df.drop(columns=dropSchedCol)
podf = podf.drop(columns=dropPoCol)
orderdf = orderdf.drop(columns=dropOrdCol)
# rename to more defined user names
df = df.rename(
    columns={
        oldCust: cust,
        oldPoNum: custPo,
        oldPartNum: partNum,
        oldRelType: relType,
        oldCustQty: custQty,
        oldDueDate: dueDate,
        oldWorkCen: workCen,
    }
)
df.index.name = jobNum

dfColumns = [
    ordDate,
    cust,
    custPo,
    partNum,
    relType,
    custQty,
    dueDate,
    workCode,
    workCen,
    estShipPrev2,
    estShipPrev,
    estShip,
    notes,
    poLink,
]

podf = podf.rename(
    columns={
        oldPoNum: colePo,
        oldPoDate: poDate,
        oldPoPartNum: poPartNum,
        oldPoPartDesc: poPartDesc,
        oldPoTotOrd: totPoQty,
        oldPoVen: poVen,
        oldPoDueDate: poDueDate,
        oldPoQtyOrd: poQty,
        oldPoQtyRec: poQtyRec,
        oldPoQtyRej: poQtyRej,
    }
)
podf.index.name = jobNum
# update the quantity type to integer
# and group blanket orders to be one job
df[custQty] = df[custQty].str.replace(",", "").astype("Int64")
qty = df[custQty].groupby(jobNum).sum()
df = df[df[partNum].notna()]
df.loc[:, custQty] = qty

podf[[totPoQty, poQty, poQtyRec, poQtyRej]] = podf[
    [totPoQty, poQty, poQtyRec, poQtyRej]
].astype("Int64")
df[estShipPrev2] = pd.NA
df[estShipPrev2] = df[estShipPrev2].astype("datetime64[ns]")
df[estShipPrev] = pd.NA
df[estShipPrev] = df[estShipPrev].astype("datetime64[ns]")
df[estShip] = pd.NA
df[estShip] = df[estShip].astype("datetime64[ns]")
df[notes] = pd.NA
df[notes] = df[notes].astype("string")
df[poLink] = pd.NA
df[poLink] = df[poLink].astype("string")
df[ordDate] = pd.NA
df[ordDate] = df[ordDate].astype("string")
df[workCode] = pd.NA
df[workCode] = df[workCode].astype("string")
podf[poLink] = pd.NA
podf[poLink] = podf[poLink].astype("string")

# if there is a priorityList.xlsx file to use
if os.path.exists(originalFile):
    oldDf = pd.read_excel(
        originalFile,
        sheet_name="Priority List",
        index_col=jobNum,
        dtype="string",
        parse_dates=[estShip, estShipPrev, estShipPrev2],
    )
    oldDf = oldDf.drop(columns=dropOldCol)
    oldDf = oldDf.drop(index=oldDf.index.difference(df.index))
    for index, value in oldDf[estShipPrev2].items():
        df.loc[index, estShipPrev2] = value
    for index, value in oldDf[estShipPrev].items():
        df.loc[index, estShipPrev] = value
    for index, value in oldDf[estShip].items():
        df.loc[index, estShip] = value
    for index, value in oldDf[notes].items():
        df.loc[index, notes] = value
else:
    print(
        "Warning: there is no priorityList.xlsx file or a Back Order Priority List folder file to take ship dates from!"
    )
    print("A new file will be generated with an empty ship date section")
    print("If this was an accident then please find a backup of the file")

orderdf = orderdf.drop(index=orderdf.index.difference(df.index))
podf = podf.drop(index=podf.index.difference(df.index))

for index, value in orderdf[oldWorkCode].items():
    df.loc[index, workCode] = value
for index, value in orderdf[oldOrdDate].items():
    df.loc[index, ordDate] = value

df = df.reindex(columns=dfColumns)

df = df[df[custPo] != "COATING ONLY"]
df = df[df[custPo] != "coating only"]

df = df.sort_index()
podf = podf.sort_index()

priWorksheet = "Priority List"
poWorksheet = "PO List"

writer = pd.ExcelWriter(newFile)
df.to_excel(writer, sheet_name=priWorksheet)
podf.to_excel(writer, sheet_name=poWorksheet)
writer.close()

wb = load_workbook(filename=newFile)
ws = wb.active
wspo = wb[poWorksheet]

bd = Side(border_style="medium", color="000000")


def toLetter(col, df):
    """get an excel letter for the name of the dataframe column
    note that this will be a bug if there is more than one column
    with the same name i.e. the column is a slice"""
    return get_column_letter(df.columns.get_loc(col) + 2)


def cellSize(cell):
    """letterSize is the width of one letter in excel
        edgeSize is the width of the spaceing at the edge of the cell
        values found using a matrix equation from real values found on excel
        EX: for Calibri at 11 font size with assumtion that equation is
    (excel column width) = letterSize * (number of characters) + edgeSize * (two edges per column)
            [ 2.86 ] = [ 2 2 ] * [ letterSize ]
            [ 4.14 ]   [ 3 2 ] * [ edgeSize   ]
        the matrix was used to solve the two font and edge size variables"""
    value = cell.value
    data_type = cell.data_type
    letterSize = 1.28
    edgeSize = 0.15
    if value is None:
        numChar = 1.0
    elif data_type == "d":
        numChar = 10.0
    elif data_type == "f":
        numChar = 8.0
    else:
        numChar = float(len(str(value)))
    if numChar > 25.0:
        numChar = 25.0
    cellSize = (letterSize * numChar) + (edgeSize * 2.0)
    return cellSize


dueDateCells = "{col}{rowBeg}:{col}{rowEnd}".format(
    col=toLetter(dueDate, df),
    rowBeg=2,
    rowEnd=len(df.index) + 1,
)

estShipCells = "{col}{rowBeg}:{col}{rowEnd}".format(
    col=toLetter(estShip, df),
    rowBeg=2,
    rowEnd=len(df.index) + 1,
)

estShipPrevCells = "{col}{rowBeg}:{col}{rowEnd}".format(
    col=toLetter(estShipPrev, df),
    rowBeg=2,
    rowEnd=len(df.index) + 1,
)

estShipPrevCells2 = "{col}{rowBeg}:{col}{rowEnd}".format(
    col=toLetter(estShipPrev2, df),
    rowBeg=2,
    rowEnd=len(df.index) + 1,
)

poClosedCells = "{colBeg}{rowBeg}:{colEnd}{rowEnd}".format(
    colBeg=toLetter(colePo, podf),
    rowBeg=2,
    colEnd=toLetter(poQtyRej, podf),
    rowEnd=len(podf.index) + 1,
)

lateFill = PatternFill(start_color="EE1111", end_color="EE1111", fill_type="solid")
poClosed = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
estShipAndNotesFill = PatternFill(
    start_color="FFCC99", end_color="FFCC99", fill_type="solid"
)
linkFill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

"""
lettCustPo = toLetter(custPo, df)
lettCust = toLetter(cust, df)
lettPartNum = toLetter(partNum, df)
lettWorkCen = toLetter(workCen, df)
lettNotes = toLetter(notes, df)
for name, group in df.groupby("Job Num"):
    row = df.index.get_loc(name)
    excelRowBeg = 0
    excelRowEnd = 0
    if type(row) == slice:
        print(f"{lettCust}{row.start+2}:{lettCust}{row.stop+1}", end=" ")
        print(f"{lettCustPo}{row.start+2}:{lettCust}{row.stop+1}", end=" ")
        print(f"{lettPartNum}{row.start+2}:{lettCust}{row.stop+1}", end=" ")
        print(f"{lettWorkCen}{row.start+2}:{lettCust}{row.stop+1}", end=" ")
        print(f"{lettNotes}{row.start+2}:{lettCust}{row.stop+1}")
        """

# creating hyperlink for po on each job
for name, group in df.groupby(jobNum):
    if name in podf.index:
        row = df.index.get_loc(name)
        poRow = podf.index.get_loc(name)
        if type(row) == slice:
            row = row.start + 2
        else:
            row = row + 2
        if type(poRow) == slice:
            poRow = poRow.start + 2
        else:
            poRow = poRow + 2
        link = f"""=HYPERLINK("#'{poWorksheet}'!A{poRow}:{toLetter(poQtyRej, podf)}{poRow}","------------>")"""
        valueLoc = f"{toLetter(poLink, df)}{row}"
        ws[valueLoc].fill = linkFill
        ws[valueLoc] = link

# creating hyperlink for job on each po
for name, group in podf.groupby(jobNum):
    if name in df.index:
        row = podf.index.get_loc(name)
        jobRow = df.index.get_loc(name)
        if type(row) == slice:
            row = row.start + 2
        else:
            row = row + 2
        if type(jobRow) == slice:
            jobRow = jobRow.start + 2
        else:
            jobRow = jobRow + 2
        link = f"""=HYPERLINK("#'{priWorksheet}'!A{jobRow}:{toLetter(notes, df)}{jobRow}","------------>")"""
        valueLoc = f"{toLetter(poLink, podf)}{row}"
        wspo[valueLoc].fill = linkFill
        wspo[valueLoc] = link

ws.conditional_formatting.add(
    dueDateCells,
    CellIsRule(operator="lessThanOrEqual", formula=["TODAY()+7"], fill=lateFill),
)

ws.conditional_formatting.add(
    estShipCells,
    FormulaRule(
        formula=[
            f"AND({toLetter(estShip, df)}2 <= TODAY(),NOT(ISBLANK({toLetter(estShip, df)}2)))"
        ],
        fill=lateFill,
    ),
)

ws.conditional_formatting.add(
    estShipPrevCells,
    FormulaRule(
        formula=[
            f"AND({toLetter(estShipPrev, df)}2 <= TODAY(),NOT(ISBLANK({toLetter(estShipPrev, df)}2)))"
        ],
        fill=lateFill,
    ),
)

ws.conditional_formatting.add(
    estShipPrevCells2,
    FormulaRule(
        formula=[
            f"AND({toLetter(estShipPrev2, df)}2 <= TODAY(),NOT(ISBLANK({toLetter(estShipPrev2, df)}2)))"
        ],
        fill=lateFill,
    ),
)

wspo.conditional_formatting.add(
    poClosedCells,
    FormulaRule(
        formula=[f"${toLetter(poQty, podf)}2 = ${toLetter(poQtyRec, podf)}2"],
        fill=poClosed,
    ),
)

coln = 0
for col in ws.iter_cols(
    min_row=1, max_col=(len(df.columns) + 1), max_row=(len(df.index) + 1)
):
    for cell in col:
        cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        if (
            (coln == df.columns.get_loc("Due Date") + 1)
            | (coln == df.columns.get_loc(estShip) + 1)
            | (coln == df.columns.get_loc(estShipPrev) + 1)
            | (coln == df.columns.get_loc(estShipPrev2) + 1)
        ):
            cell.number_format = "mm/dd/yyyy"
        elif (coln == df.columns.get_loc(estShip) + 1) | (
            coln == df.columns.get_loc(notes) + 1
        ):
            cell.fill = estShipAndNotesFill
    coln = coln + 1

coln = 0
for col in wspo.iter_cols(
    min_row=1, max_col=(len(podf.columns) + 1), max_row=(len(podf.index) + 1)
):
    for cell in col:
        cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        if (coln == podf.columns.get_loc("PO Due Date") + 1) | (
            coln == podf.columns.get_loc("Cole PO Date") + 1
        ):
            cell.number_format = "mm/dd/yyyy"
    coln = coln + 1

# auto format column widths for both worksheets
for column_cells in ws.columns:
    length = max((cellSize(cell)) for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

for column_cells in wspo.columns:
    length = max((cellSize(cell)) for cell in column_cells)
    wspo.column_dimensions[get_column_letter(column_cells[0].column)].width = length

lastColumn = get_column_letter(len(df.columns) + 1)
lastRow = len(df.index) + 1

ws.freeze_panes = ws["A2"]
wspo.freeze_panes = wspo["A2"]

filterRangeMain = "A1:" + toLetter(notes, df) + str(lastRow)
filterRangePo = "A1:" + toLetter(poQtyRej, podf) + str(lastRow)
ws.auto_filter.ref = filterRangeMain
wspo.auto_filter.ref = filterRangePo

# save a version of the final file locally
wb.save(newFile)


# take the locally generated file and apply it to the quote scans folder
final_source = os.path.relpath(f"./{newFile}")
final_dest = os.path.abspath(f"{quote_scan_folder}{newFile}")
try:
    shutil.move(final_source, final_dest)
except FileNotFoundError:
    print(f"Error: Source file {final_source} not found.")
except Exception as e:
    print(f"An error occurred: {e}")
